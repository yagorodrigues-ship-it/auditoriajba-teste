import streamlit as st
import pandas as pd
import datetime
import io
import re
import time
import gspread
from google.oauth2.service_account import Credentials
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Auditoria & Inventário - JBA (TESTE DRIVE)", layout="wide")

# --- CONEXÃO COM O GOOGLE SHEETS VIA SERVICE ACCOUNT ---
@st.cache_resource
def conectar_gspread():
    try:
        scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        info_dict = dict(st.secrets["connections"]["gsheets"])
        creds = Credentials.from_service_account_info(info_dict, scopes=scope)
        gc = gspread.authorize(creds)
        url_planilha = st.secrets["gconfigs"]["spreadsheet_url"]
        return gc.open_by_url(url_planilha)
    except Exception as e:
        st.error(f"❌ Erro de Autenticação com o Google Sheets: {e}")
        return None

def ler_aba(nome_aba):
    """Lê uma aba do Google Sheets padronizando os cabeçalhos."""
    try:
        sh = conectar_gspread()
        if sh:
            ws = sh.worksheet(nome_aba)
            dados = ws.get_all_records()
            if dados:
                df = pd.DataFrame(dados).fillna("")
                df.columns = [str(c).replace('\xa0', ' ').strip().lower() for c in df.columns]
                return df
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()

def sincronizar_upsert_inventarios(df_novos_inv):
    """Atualiza o registro do inventário na planilha pelo ID sem duplicar linhas."""
    try:
        sh = conectar_gspread()
        if sh and not df_novos_inv.empty:
            ws = sh.worksheet("inventarios")
            df_existente = ler_aba("inventarios")
            
            df_novos_inv['id_clean'] = df_novos_inv['id'].astype(str).str.replace('#', '', regex=False).str.strip()
            
            if not df_existente.empty:
                df_existente['id_clean'] = df_existente['id'].astype(str).str.replace('#', '', regex=False).str.strip()
                ids_novos = df_novos_inv['id_clean'].tolist()
                df_mantidos = df_existente[~df_existente['id_clean'].isin(ids_novos)]
                df_final = pd.concat([df_mantidos, df_novos_inv], ignore_index=True)
            else:
                df_final = df_novos_inv

            if 'id_clean' in df_final.columns:
                df_final = df_final.drop(columns=['id_clean'])
                
            df_final = df_final.astype(str)
            ws.clear()
            ws.update([df_final.columns.values.tolist()] + df_final.values.tolist())
            st.cache_data.clear()
            return True
        return False
    except Exception as e:
        st.error(f"❌ Erro ao sincronizar inventários: {e}")
        return False

def salvar_lote_aba(nome_aba, novos_dados_df):
    """Adiciona os registros ao final da aba mantendo o histórico existente."""
    try:
        sh = conectar_gspread()
        if sh and not novos_dados_df.empty:
            ws = sh.worksheet(nome_aba)
            novos_dados_df.columns = [str(c).replace('\xa0', ' ').strip().lower() for c in novos_dados_df.columns]
            
            df_existente = ler_aba(nome_aba)
            if not df_existente.empty:
                df_final = pd.concat([df_existente, novos_dados_df], ignore_index=True)
            else:
                df_final = novos_dados_df
                
            df_final = df_final.astype(str)
            ws.clear()
            ws.update([df_final.columns.values.tolist()] + df_final.values.tolist())
            st.cache_data.clear()
            return True
        return False
    except Exception as e:
        st.error(f"❌ Erro ao salvar na aba '{nome_aba}': {e}")
        return False

def atualizar_aba_completa(nome_aba, df_completo):
    """Substitui o conteúdo de uma aba garantindo a limpeza das linhas antigas."""
    try:
        sh = conectar_gspread()
        if sh:
            ws = sh.worksheet(nome_aba)
            if not df_completo.empty:
                df_completo.columns = [str(c).replace('\xa0', ' ').strip().lower() for c in df_completo.columns]
                df_completo = df_completo.astype(str)
                ws.clear()
                ws.update([df_completo.columns.values.tolist()] + df_completo.values.tolist())
            else:
                ws.clear()
            st.cache_data.clear()
            return True
        return False
    except Exception as e:
        st.error(f"❌ Erro ao atualizar aba '{nome_aba}': {e}")
        return False

# --- LISTA OFICIAL E UNIFICADA DE ESTOQUES JBA ---
LISTA_ESTOQUES_FIXA = [
    {"id": "1077", "desc": "JBA - CLASSE D"}, {"id": "1078", "desc": "JBA - COPA E COZINHA"},
    {"id": "1080", "desc": "JBA - DADOS - CLIENTE"}, {"id": "1082", "desc": "JBA - VIVO VITA - CLIENTE"},
    {"id": "1084", "desc": "JBA - EPI-EPC"}, {"id": "1086", "desc": "JBA - EQUIPAMENTOS"},
    {"id": "1088", "desc": "JBA - FERRAMENTAL"}, {"id": "1089", "desc": "JBA - KIT FERRAMENTAL CONTRATACOES"},
    {"id": "1090", "desc": "JBA - FERRAMENTAS DE CANTEIRO"}, {"id": "1104", "desc": "JBA - MATERIAL DE ESCRITORIO - SUPRIMENTOS DE INFORMATICA"},
    {"id": "1106", "desc": "JBA - MOBILIARIO"}, {"id": "1113", "desc": "1385 - MANUTENCAO JBA - CLIENTE"},
    {"id": "1118", "desc": "JBA - PROPRIO GERAL"}, {"id": "1122", "desc": "JBA - GRAND OBRAS IMPLANTACAO"},
    {"id": "1140", "desc": "JBA - SPEEDY/FTTX - CLIENTE"}, {"id": "1144", "desc": "1385 - MANUTENCAO JBA CLIENTE RESERVADO"},
    {"id": "1149", "desc": "JBA - UNIFORME"}, {"id": "2183", "desc": "1071 - BOL IMPLANTANCAO JBA - CLIENTE"},
    {"id": "2185", "desc": "JBA - PROPRIO FATURA B PLANTA EXTERNA - BDI"}, {"id": "2188", "desc": "1071 - IMPLANTACAO JBA CLIENTE RESERVADO"},
    {"id": "2190", "desc": "JBA - DEPARTAMENTO T.I"}, {"id": "2194", "desc": "JBA - KITS FERRAMENTAL - DEVOLUCAO"},
    {"id": "2197", "desc": "JBA - EQUIPAMENTOS TI"}, {"id": "2641", "desc": "1259 - IMPLANTACAO JBA - MATERIAL REUTILIZACAO"},
    {"id": "2643", "desc": "1724 - MANUTENCAO JBA - MATERIAL REUTILIZACAO"}, {"id": "2725", "desc": "JBA - RESERVA TIM"},
    {"id": "2983", "desc": "JBA - FORNECEDORES P/ MANUTENCAO - RECARGA"}, {"id": "3193", "desc": "JBA - PROPRIO MATERIAL REAPROVEITAVEL"},
    {"id": "3395", "desc": "LPA - FTTX - CLIENTE"}, {"id": "3484", "desc": "JBA - CELULARES DEFEITO"},
    {"id": "3546", "desc": "JBA - CELULARES"}
]
MAPA_ESTOQUES_DESC = {item['id']: item['desc'] for item in LISTA_ESTOQUES_FIXA}

def limpar_cache_aplicacao():
    st.cache_data.clear()

def limpar_documento(doc):
    return str(doc).strip().replace(".", "").replace("-", "").replace("/", "")

def extrair_id_estoque_do_nome(nome_inventario):
    numeros = re.findall(r'\b\d{4}\b', str(nome_inventario))
    for num in numeros:
        if num in MAPA_ESTOQUES_DESC:
            return num
    return ""

def converter_para_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Relatorio')
    return output.getvalue()

def gerar_relatorio_consolidado_excel(df_contagens, lista_estoques_ref):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        linhas_resumo = []
        for est in lista_estoques_ref:
            est_id = str(est["id"]).strip()
            df_est = df_contagens[df_contagens['id_estoque'].astype(str).str.strip() == est_id] if not df_contagens.empty else pd.DataFrame()
            
            total_itens = len(df_est)
            acertos = len(df_est[pd.to_numeric(df_est['diferenca'], errors='coerce') == 0]) if total_itens > 0 else 0
            erros = len(df_est[pd.to_numeric(df_est['diferenca'], errors='coerce') != 0]) if total_itens > 0 else 0
            qtd_segunda = len(df_est[df_est['fase_contagem'] == '2a Contagem']) if 'fase_contagem' in df_est.columns and total_itens > 0 else 0
            acuracidade_pct = (acertos / total_itens) if total_itens > 0 else 0.0
            u_data_est = str(df_est['data_hora'].max()) if not df_est.empty and 'data_hora' in df_est.columns else "—"
            
            obs_erros = []
            if not df_est.empty and 'observacao' in df_est.columns:
                obs_validas = df_est[df_est['observacao'].fillna('').astype(str).str.strip() != '']['observacao'].unique().tolist()
                if obs_validas: obs_erros.append("; ".join(map(str, obs_validas)))
            if erros > 0 and not df_est.empty:
                erros_cods = df_est[pd.to_numeric(df_est['diferenca'], errors='coerce') != 0]['cod_produto'].unique().tolist()
                obs_erros.append(f"Divergências: {', '.join(map(str, erros_cods[:5]))}" + ("..." if len(erros_cods) > 5 else ""))
            
            linhas_resumo.append({
                "Id. Estoq. Físico": est_id, "Desc. Estoque Físico": est["desc"],
                "Total de itens": total_itens, "Acertos": acertos, "Erros": erros,
                "2º Contagem": qtd_segunda, "Acuracidade": acuracidade_pct,
                "Última Contagem": u_data_est, "Obs": " | ".join(obs_erros) if obs_erros else "Nenhuma observação"
            })
            
        df_resumo = pd.DataFrame(linhas_resumo)
        df_resumo.to_excel(writer, index=False, sheet_name='Acuracidade')
        
        worksheet = writer.sheets['Acuracidade']
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        for row in range(2, len(df_resumo) + 2):
            worksheet[f'G{row}'].number_format = '0.00%'
            cell_erros = worksheet[f'E{row}']
            if cell_erros.value and int(cell_erros.value) > 0: cell_erros.fill = red_fill
            elif cell_erros.value == 0 and worksheet[f'C{row}'].value and int(worksheet[f'C{row}'].value) > 0: cell_erros.fill = green_fill

        if not df_contagens.empty:
            for est_id, group in df_contagens.groupby('id_estoque'):
                nome_aba = str(est_id).strip()[:31]
                if nome_aba:
                    group_limpo = group.drop(columns=['id'], errors='ignore')
                    group_limpo.to_excel(writer, index=False, sheet_name=nome_aba)
                    ws_est = writer.sheets[nome_aba]
                    rule_error = CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill)
                    rule_ok = CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=green_fill)
                    ws_est.conditional_formatting.add(f"I2:I{len(group_limpo)+1}", rule_error)
                    ws_est.conditional_formatting.add(f"I2:I{len(group_limpo)+1}", rule_ok)

    return output.getvalue()

# --- ESTILOS CSS ---
st.markdown("""
    <style>
    div.stButton > button:first-child[kind="primary"] { background-color: #d35400; border-color: #d35400; color: white; font-weight: bold; }
    div.stButton > button:first-child[kind="primary"]:hover { background-color: #e67e22; border-color: #e67e22; color: white; }
    .card-lateral { background-color: #1a233a; padding: 10px; border-radius: 8px; margin-bottom: 8px; border-left: 4px solid #2563eb; }
    .card-lateral-titulo { color: #93c5fd; font-size: 11px; font-weight: bold; }
    .card-lateral-valor { color: white; font-size: 22px; font-weight: bold; }
    .caixa-descricao-produto { background-color: #ebf8ff; border-left: 5px solid #3182ce; padding: 12px 16px; border-radius: 6px; font-size: 16px; font-weight: bold; color: #1a365d; margin: 10px 0; }
    </style>
""", unsafe_allow_html=True)

# --- INICIALIZAÇÃO DE ESTADOS EM MEMÓRIA RAM ---
if 'logged_in' not in st.session_state: st.session_state.logged_in = False
if 'operador' not in st.session_state: st.session_state.operador = ""
if 'perfil_usuario' not in st.session_state: st.session_state.perfil_usuario = "Almoxarife"
if 'tela_acesso' not in st.session_state: st.session_state.tela_acesso = "login"
if 'contador_reset' not in st.session_state: st.session_state.contador_reset = 0
if 'contador_reset_sup' not in st.session_state: st.session_state.contador_reset_sup = 0
if 'bases_supervisor_por_inv' not in st.session_state: st.session_state.bases_supervisor_por_inv = {}
if 'pagina_historico' not in st.session_state: st.session_state.pagina_historico = 1

if 'buffer_ram_inventarios' not in st.session_state: st.session_state.buffer_ram_inventarios = []
if 'buffer_ram_contagens' not in st.session_state: st.session_state.buffer_ram_contagens = []
if 'buffer_ram_bases' not in st.session_state: st.session_state.buffer_ram_bases = []

def sincronizar_ram_com_banco():
    tot_inv = len(st.session_state.buffer_ram_inventarios)
    tot_cnt = len(st.session_state.buffer_ram_contagens)
    tot_bas = len(st.session_state.buffer_ram_bases)
    
    sucesso_total = True
    
    if tot_inv > 0:
        if sincronizar_upsert_inventarios(pd.DataFrame(st.session_state.buffer_ram_inventarios)):
            st.session_state.buffer_ram_inventarios = []
        else: sucesso_total = False
        
    if tot_cnt > 0:
        df_cnts_ram = pd.DataFrame(st.session_state.buffer_ram_contagens)
        if salvar_lote_aba("contagens", df_cnts_ram):
            df_ultimas = ler_aba("ultima_contagem_estoques")
            novas_datas = []
            for item in st.session_state.buffer_ram_contagens:
                if item.get('id_estoque'):
                    novas_datas.append({'id_estoque': str(item['id_estoque']), 'ultima_data': str(item['data_hora'])})
            if novas_datas:
                df_novas_dt = pd.DataFrame(novas_datas)
                if not df_ultimas.empty:
                    df_ultimas = pd.concat([df_ultimas, df_novas_dt], ignore_index=True).drop_duplicates(subset=['id_estoque'], keep='last')
                else:
                    df_ultimas = df_novas_dt
                atualizar_aba_completa("ultima_contagem_estoques", df_ultimas)
            st.session_state.buffer_ram_contagens = []
        else: sucesso_total = False

    if tot_bas > 0:
        if salvar_lote_aba("itens_base_inventario", pd.DataFrame(st.session_state.buffer_ram_bases)):
            st.session_state.buffer_ram_bases = []
        else: sucesso_total = False

    limpar_cache_aplicacao()
    return (tot_inv + tot_cnt + tot_bas) if sucesso_total else 0

# --- TELA DE LOGIN CENTRALIZADA ---
if not st.session_state.logged_in:
    col_vaz1, col_central, col_vaz2 = st.columns([1, 1.2, 1])
    with col_central:
        if st.session_state.tela_acesso == "login":
            st.title("🔒 Acesso ao Sistema JBA (Drive)")
            with st.form("login_form"):
                identificador = st.text_input("CPF ou E-mail")
                senha = st.text_input("Senha", type="password")
                if st.form_submit_button("Entrar no Sistema", type="primary", use_container_width=True):
                    id_limpo = identificador.strip()
                    doc_limpo = limpar_documento(id_limpo)
                    
                    df_usrs = ler_aba("usuarios")
                    if not df_usrs.empty:
                        u_match = df_usrs[((df_usrs['email'].astype(str) == id_limpo) | (df_usrs['cpf'].astype(str) == doc_limpo)) & (df_usrs['senha'].astype(str) == senha.strip())]
                        if not u_match.empty:
                            st.session_state.logged_in = True
                            st.session_state.operador = u_match.iloc[0]['nome']
                            st.session_state.perfil_usuario = u_match.iloc[0]['perfil'] or "Almoxarife"
                            limpar_cache_aplicacao()
                            st.rerun()
                        else:
                            st.error("❌ Credenciais incorretas.")
                    else:
                        if senha == "123":
                            st.session_state.logged_in = True
                            st.session_state.operador = "Administrador Tel"
                            st.session_state.perfil_usuario = "Administrador"
                            limpar_cache_aplicacao()
                            st.rerun()
                        else:
                            st.error("❌ Base de usuários vazia no Google Sheets.")

# --- APLICAÇÃO PRINCIPAL LOGADA ---
else:
    df_inv_drive = ler_aba("inventarios")
    if not df_inv_drive.empty and 'id' in df_inv_drive.columns:
        df_inv_drive['id_clean'] = df_inv_drive['id'].astype(str).str.replace('#', '', regex=False).str.strip()
        df_inv_drive['id_num'] = pd.to_numeric(df_inv_drive['id_clean'], errors='coerce').fillna(0)
        df_inv_drive = df_inv_drive.drop_duplicates(subset=['id_clean'], keep='last')
    else:
        df_inv_drive = pd.DataFrame(columns=['id', 'nome', 'data', 'status', 'total_itens', 'acuracidade_final'])

    if st.session_state.buffer_ram_inventarios:
        df_inv_ram = pd.DataFrame(st.session_state.buffer_ram_inventarios)
        df_inv_ram['id_clean'] = df_inv_ram['id'].astype(str).str.replace('#', '', regex=False).str.strip()
        df_inventarios = pd.concat([df_inv_ram, df_inv_drive], ignore_index=True).drop_duplicates(subset=['id_clean'], keep='first')
    else:
        df_inventarios = df_inv_drive.copy()

    if not df_inventarios.empty and 'id_num' in df_inventarios.columns:
        df_inventarios = df_inventarios.sort_values(by=['data', 'id_num'], ascending=[False, False])

    df_inventarios_sup = ler_aba("inventarios_supervisor")
    eh_supervisor = (st.session_state.perfil_usuario == "Administrador") or ("admin" in st.session_state.operador.lower())

    # --- BARRA LATERAL ---
    with st.sidebar:
        st.write(f"👤 **{st.session_state.operador}** ({st.session_state.perfil_usuario})")
        
        pendentes_totais = len(st.session_state.buffer_ram_inventarios) + len(st.session_state.buffer_ram_contagens) + len(st.session_state.buffer_ram_bases)
        if pendentes_totais > 0:
            st.warning(f"⚡ **{pendentes_totais} alterações** guardadas na RAM local!")
        else:
            st.success("🟢 Memória RAM sincronizada com Drive.")

        col_s1, col_s2 = st.columns(2)
        with col_s1:
            if st.button("🔄 Atualizar", use_container_width=True):
                limpar_cache_aplicacao()
                st.rerun()
        with col_s2:
            if st.button("🚪 Sair", use_container_width=True):
                sincronizar_ram_com_banco()
                st.session_state.logged_in = False
                st.session_state.operador = ""
                limpar_cache_aplicacao()
                st.rerun()
            
        st.markdown("---")
        st.write("📁 **Seleção de Inventário**")
        
        if df_inventarios.empty or 'id' not in df_inventarios.columns:
            id_inventario_atual = None
            inventario_selected_obj = None
            id_pasta_limpo_base = ""
            st.info("Crie um inventário abaixo.")
        else:
            lista_inv = [f"{row['id']} – {row['nome']} ({row['status']})" for idx, row in df_inventarios.iterrows()]
            inventario_selected = st.selectbox("Selecione a Pasta", lista_inv, index=0, key="sb_pasta_ativa")
            id_inventario_atual = inventario_selected.split(" – ")[0]
            inventario_selected_obj = df_inventarios[df_inventarios['id'].astype(str) == id_inventario_atual].iloc[0]
            id_pasta_limpo_base = str(id_inventario_atual).replace("#", "").strip()

        # UPLOAD DA BASE
        st.write("📂 **Upload Base de Dados (.xlsx)**")
        if inventario_selected_obj is not None and str(inventario_selected_obj['status']) == "Aberto":
            uploader_key = f"func_excel_loader_{id_pasta_limpo_base}_{st.session_state.contador_reset}"
            arquivo_excel = st.file_uploader("Suba o arquivo Excel (.xlsx)", type=["xlsx"], label_visibility="collapsed", key=uploader_key)
            
            if arquivo_excel is not None and id_pasta_limpo_base:
                df_upload_temp = pd.read_excel(arquivo_excel)
                mapa_cols_upload = {str(col).replace('\xa0', ' ').strip().lower().replace(" ", "").replace(".", "").replace("ó", "o").replace("á", "a"): col for col in df_upload_temp.columns}

                col_cod, col_desc = mapa_cols_upload.get("codproduto"), mapa_cols_upload.get("descproduto")
                col_est, col_uni = mapa_cols_upload.get("descestoquefisico"), mapa_cols_upload.get("unidmedida")
                col_qtd, col_idest = mapa_cols_upload.get("qtdestoque"), mapa_cols_upload.get("idestoquefisico")
                col_lote, col_ativo = mapa_cols_upload.get("lote"), mapa_cols_upload.get("ativo")

                est_id_fallback = extrair_id_estoque_do_nome(inventario_selected_obj['nome'])
                est_desc_fallback = MAPA_ESTOQUES_DESC.get(est_id_fallback, 'ESTOQUE FÍSICO')

                if not col_cod or not col_desc:
                    st.error("❌ Planilha fora do padrão JBA.")
                else:
                    st.session_state.buffer_ram_bases = [b for b in st.session_state.buffer_ram_bases if str(b['inventario_id']) != id_pasta_limpo_base]
                    for _, r in df_upload_temp.iterrows():
                        st.session_state.buffer_ram_bases.append({
                            "inventario_id": id_pasta_limpo_base,
                            "cod_produto": str(r[col_cod]).strip() if col_cod and pd.notna(r[col_cod]) else '',
                            "desc_produto": str(r[col_desc]).strip() if col_desc and pd.notna(r[col_desc]) else '',
                            "desc_estoque_fisico": str(r[col_est]).strip() if col_est and pd.notna(r[col_est]) and str(r[col_est]).strip() != '' else est_desc_fallback,
                            "unid_medida": str(r[col_uni]).strip() if col_uni and pd.notna(r[col_uni]) else '',
                            "qtd_estoque": int(pd.to_numeric(r[col_qtd], errors='coerce') or 0) if col_qtd else 0,
                            "id_estoque_fisico": str(r[col_idest]).strip() if col_idest and pd.notna(r[col_idest]) and str(r[col_idest]).strip() != '' else est_id_fallback,
                            "lote": str(r[col_lote]).strip() if col_lote and pd.notna(r[col_lote]) and str(r[col_lote]).lower() != 'nan' else '',
                            "ativo": str(r[col_ativo]).strip() if col_ativo and pd.notna(r[col_ativo]) and str(r[col_ativo]).lower() != 'nan' else ''
                        })
                    st.toast("✅ Base Carregada com sucesso em RAM!", icon="✅")

        # CARREGA BASE ATIVA (COMBINADA DRIVE + RAM)
        df_base_drive = ler_aba("itens_base_inventario")
        if not df_base_drive.empty and 'inventario_id' in df_base_drive.columns:
            df_base_drive = df_base_drive[df_base_drive['inventario_id'].astype(str) == id_pasta_limpo_base]
        else:
            df_base_drive = pd.DataFrame()

        if st.session_state.buffer_ram_bases:
            df_base_ram = pd.DataFrame([b for b in st.session_state.buffer_ram_bases if str(b['inventario_id']) == id_pasta_limpo_base])
            base_sistema_atual = pd.concat([df_base_ram, df_base_drive], ignore_index=True)
        else:
            base_sistema_atual = df_base_drive.copy()

        if inventario_selected_obj is not None and str(inventario_selected_obj['status']) == "Aberto" and not base_sistema_atual.empty:
            st.markdown("---")
            if st.button("🚀 Salvar Base e Iniciar 1ª Contagem", type="primary", use_container_width=True):
                for inv in st.session_state.buffer_ram_inventarios:
                    if str(inv['id']) == str(id_inventario_atual):
                        inv['status'] = '1a Contagem'

                df_inv_drive = ler_aba("inventarios")
                if not df_inv_drive.empty:
                    df_inv_drive.loc[df_inv_drive['id'].astype(str) == str(id_inventario_atual), 'status'] = '1a Contagem'
                    sincronizar_upsert_inventarios(df_inv_drive)

                sincronizar_ram_com_banco()
                limpar_cache_aplicacao()
                st.success("🔒 1ª Contagem liberada.")
                st.rerun()

        # CRIAÇÃO DE NOVO INVENTÁRIO
        with st.expander("➕ Criar Novo Inventário"):
            nome_nova_pasta = st.text_input("Nome do Inventário", key="txt_novo_nome_pasta")
            if st.button("Criar Pasta Agora", type="primary", use_container_width=True):
                if not nome_nova_pasta.strip():
                    st.error("⚠️ Digite um nome para a pasta!")
                else:
                    maior_id = len(df_inventarios) + 1
                    novo_id_str = f"#{maior_id}"

                    st.session_state.buffer_ram_inventarios.insert(0, {
                        "id": novo_id_str,
                        "nome": nome_nova_pasta.strip(),
                        "data": datetime.date.today().strftime("%Y-%m-%d"),
                        "status": "Aberto",
                        "total_itens": 0,
                        "acuracidade_final": "0%"
                    })
                    st.toast(f"✅ Pasta {novo_id_str} criada na memória local!", icon="🎉")
                    time.sleep(0.3)
                    st.rerun()

        # FECHAMENTO DO INVENTÁRIO
        pode_fechar, itens_faltantes, itens_pendentes_2a = False, [], []
        if inventario_selected_obj is not None and str(inventario_selected_obj['status']) in ["1a Contagem", "2a Contagem"] and not base_sistema_atual.empty:
            df_cnts_todas = ler_aba("contagens")
            df_cnts_pasta = df_cnts_todas[df_cnts_todas['inventario_id'].astype(str) == id_pasta_limpo_base] if not df_cnts_todas.empty else pd.DataFrame()

            if str(inventario_selected_obj['status']) == "1a Contagem":
                set_contados_triade = set()
                if not df_cnts_pasta.empty:
                    for _, r in df_cnts_pasta.iterrows():
                        set_contados_triade.add(f"{str(r.get('cod_produto', '')).upper().strip()}_{str(r.get('lote', '')).upper().strip()}_{str(r.get('ativo', '')).upper().strip()}")

                for r_ram in st.session_state.buffer_ram_contagens:
                    if str(r_ram['inventario_id']) == id_pasta_limpo_base:
                        set_contados_triade.add(f"{str(r_ram['cod_produto']).upper().strip()}_{str(r_ram['lote']).upper().strip()}_{str(r_ram['ativo']).upper().strip()}")

                for _, r_b in base_sistema_atual.iterrows():
                    c_b = str(r_b.get('cod_produto', '')).upper().strip()
                    l_b = str(r_b.get('lote', '')).upper().strip() if pd.notna(r_b.get('lote')) and str(r_b.get('lote')).lower() != 'nan' else ""
                    a_b = str(r_b.get('ativo', '')).upper().strip() if pd.notna(r_b.get('ativo')) and str(r_b.get('ativo')).lower() != 'nan' else ""
                    if f"{c_b}_{l_b}_{a_b}" not in set_contados_triade: itens_faltantes.append(c_b)
                if len(itens_faltantes) == 0: pode_fechar = True

            elif str(inventario_selected_obj['status']) == "2a Contagem":
                bips_novos_2a = set()
                for r_ram in st.session_state.buffer_ram_contagens:
                    if str(r_ram['inventario_id']) == id_pasta_limpo_base and str(r_ram.get('fase_contagem')) in ['2a Contagem', '2a Contagem Concluida']:
                        bips_novos_2a.add(str(r_ram['cod_produto']).upper().strip())

                if not df_cnts_pasta.empty:
                    itens_liberados = df_cnts_pasta[df_cnts_pasta['fase_contagem'] == '2a Contagem']['cod_produto'].astype(str).str.upper().str.strip().tolist()
                    itens_pendentes_2a = [cod for cod in itens_liberados if cod not in bips_novos_2a]
                
                if len(itens_pendentes_2a) == 0: pode_fechar = True

            st.markdown("---")
            def fechar_e_preservar_historico(id_inv, id_limpo):
                sincronizar_ram_com_banco()
                
                df_cnts = ler_aba("contagens")
                df_f = df_cnts[df_cnts['inventario_id'].astype(str) == id_limpo] if not df_cnts.empty else pd.DataFrame()
                tot = len(df_f)
                acertos = len(df_f[pd.to_numeric(df_f['diferenca'], errors='coerce') == 0]) if tot > 0 else 0
                pct_acu = f"{(acertos / tot)*100:.1f}%" if tot > 0 else "0%"
                
                df_inv_up = ler_aba("inventarios")
                if not df_inv_up.empty:
                    id_limpo_str = str(id_inv).replace('#', '').strip()
                    mascara = df_inv_up['id'].astype(str).str.replace('#', '').str.strip() == id_limpo_str
                    df_inv_up.loc[mascara, 'status'] = 'Fechado'
                    df_inv_up.loc[mascara, 'total_itens'] = tot
                    df_inv_up.loc[mascara, 'acuracidade_final'] = pct_acu
                    sincronizar_upsert_inventarios(df_inv_up)
                limpar_cache_aplicacao()

            if pode_fechar:
                if st.button("🔒 Fechar Inventário (100% Concluído)", use_container_width=True, type="primary"):
                    fechar_e_preservar_historico(id_inventario_atual, id_pasta_limpo_base)
                    st.success("✅ Inventário encerrado e arquivado!")
                    st.rerun()
            else:
                qtd_f = len(itens_faltantes) if str(inventario_selected_obj['status']) == "1a Contagem" else len(itens_pendentes_2a)
                st.warning(f"⏳ Faltam **{qtd_f}** itens para concluir.")
                if eh_supervisor:
                    if st.button("🚨 Forçar Fechamento Incompleto (ADMIN)", use_container_width=True):
                        fechar_e_preservar_historico(id_inventario_atual, id_pasta_limpo_base)
                        st.success("✅ Inventário encerrado!")
                        st.rerun()

        # KPIs SIDEBAR
        total_itens_base = len(base_sistema_atual) if not base_sistema_atual.empty else 0
        df_cnt_cnt = ler_aba("contagens")
        df_cnt_pasta = df_cnt_cnt[df_cnt_cnt['inventario_id'].astype(str) == id_pasta_limpo_base] if not df_cnt_cnt.empty else pd.DataFrame()
        total_contados_cnt = len(df_cnt_pasta) + len([c for c in st.session_state.buffer_ram_contagens if str(c['inventario_id']) == id_pasta_limpo_base])
        
        st.markdown(f'<div class="card-lateral"><div class="card-lateral-titulo">📋 ITENS NA BASE</div><div class="card-lateral-valor">{total_itens_base}</div></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="card-lateral"><div class="card-lateral-titulo">✅ LANÇAMENTOS</div><div class="card-lateral-valor">{total_contados_cnt}</div></div>', unsafe_allow_html=True)

        st.markdown("---")
        if st.button("🚀 Sincronizar Agora com o Drive", type="primary", use_container_width=True):
            if pendentes_totais > 0:
                qtd = sincronizar_ram_com_banco()
                if qtd > 0:
                    st.success(f"✅ {qtd} alterações enviadas para a planilha do Drive!")
                else:
                    st.error("❌ Falha na gravação. Verifique as permissões da planilha.")
            else:
                st.info("ℹ️ Nenhuma alteração pendente em RAM.")
            st.rerun()

    # --- DECLARAÇÃO DAS ABAS PRINCIPAIS ---
    lista_abas = ["🔍 Contar Item (RAM Modo Rápido)", "📊 Lançamentos & Base", "📈 Desempenho & Acuracidade", "📁 Histórico Geral"]
    if eh_supervisor: lista_abas.append("⚙️ Gestão ADM")
    
    abas_objs = st.tabs(lista_abas)
    aba_contar, aba_lancamentos, aba_desempenho, aba_historico = abas_objs[0], abas_objs[1], abas_objs[2], abas_objs[3]
    aba_adm = abas_objs[4] if eh_supervisor else None

    # --- ABA 1: CONTAR ITEM (RAM MODO RÁPIDO) ---
    with aba_contar:
        if pendentes_totais > 0:
            c_top1, c_top2 = st.columns([3, 1])
            c_top1.info(f"⚡ **Modo Ultra Rápido Ativo:** Você tem **{pendentes_totais} bips/dados** guardados na RAM local.")
            with c_top2:
                df_ram_temp = pd.DataFrame(st.session_state.buffer_ram_contagens)
                st.download_button("📥 Backup RAM (.xlsx)", converter_para_excel(df_ram_temp), file_name=f"backup_ram_pasta_{id_pasta_limpo_base}.xlsx")

        if not id_inventario_atual or base_sistema_atual.empty:
            st.warning("⚠️ Selecione um inventário ativo e carregue a base na barra lateral.")
        elif str(inventario_selected_obj['status']) == "Aberto": st.warning("⚠️ Inventário em configuração. Libere a 1ª Contagem na barra lateral.")
        elif str(inventario_selected_obj['status']) == "Fechado": st.error("🔒 Inventário Fechado. Selecione um inventário ativo.")
        elif pode_fechar and str(inventario_selected_obj['status']) in ["1a Contagem", "2a Contagem"]:
            st.success("🎉 **100% dos itens desta fase já foram contados!** Você pode fechar o inventário na barra lateral.")
        else:
            codigo_input = st.text_input("💻 Bipar ou Digitar Código do Produto", value="", placeholder="Bipe a etiqueta aqui...", key=f"bip_{st.session_state.contador_reset}")
            if codigo_input:
                codigo_rastreio = str(codigo_input).upper().strip().split(" - ")[-1]
                matches_codigo = base_sistema_atual[base_sistema_atual['cod_produto'].astype(str).str.upper().str.strip() == codigo_rastreio]
                if matches_codigo.empty: st.error("❌ Código não cadastrado na planilha base!")
                else:
                    df_cnts_exist = ler_aba("contagens")
                    df_cnts_exist_pasta = df_cnts_exist[df_cnts_exist['inventario_id'].astype(str) == id_pasta_limpo_base] if not df_cnts_exist.empty else pd.DataFrame()
                    
                    status_pasta_atual = str(inventario_selected_obj['status'])

                    if status_pasta_atual == "2a Contagem":
                        if not df_cnts_exist_pasta.empty:
                            itens_liberados_2a = df_cnts_exist_pasta[df_cnts_exist_pasta['fase_contagem'] == '2a Contagem']['cod_produto'].astype(str).str.upper().str.strip().tolist()
                            if codigo_rastreio not in itens_liberados_2a:
                                st.warning(f"⚠️ O item {codigo_rastreio} já teve a 1ª contagem ok e NÃO está pendente de 2ª contagem.")
                                pode_exibir_form = False
                            else:
                                matches_para_usar = matches_codigo
                                pode_exibir_form = True
                        else:
                            matches_para_usar = matches_codigo
                            pode_exibir_form = True
                    else:
                        set_ja_contados = set()
                        if not df_cnts_exist_pasta.empty:
                            for _, r in df_cnts_exist_pasta[df_cnts_exist_pasta['cod_produto'].astype(str).str.upper().str.strip() == codigo_rastreio].iterrows():
                                set_ja_contados.add(f"{str(r.get('lote', '')).strip().upper()}_{str(r.get('ativo', '')).strip().upper()}")
                        
                        for r_ram in st.session_state.buffer_ram_contagens:
                            if str(r_ram['cod_produto']).upper().strip() == codigo_rastreio:
                                set_ja_contados.add(f"{str(r_ram['lote']).strip().upper()}_{str(r_ram['ativo']).strip().upper()}")

                        matches_pendentes = [row_m for _, row_m in matches_codigo.iterrows() if f"{str(row_m.get('lote', '')).strip().upper()}_{str(row_m.get('ativo', '')).strip().upper()}" not in set_ja_contados]
                        if not matches_pendentes:
                            st.success(f"🎉 **Todos os ativos/lotes do produto {codigo_rastreio} já foram contabilizados!**")
                            pode_exibir_form = False
                        else:
                            matches_para_usar = pd.DataFrame(matches_pendentes)
                            pode_exibir_form = True

                    if pode_exibir_form:
                        if len(matches_para_usar) > 1:
                            opcoes_item = [f"Ativo: {str(r.get('ativo', '')).strip() or 'Sem Ativo'} | Lote: {str(r.get('lote', '')).strip() or 'Sem Lote'}" for _, r in matches_para_usar.iterrows()]
                            item_sel_opcao = st.selectbox("Escolha o item específico PENDENTE:", opcoes_item)
                            idx_match = opcoes_item.index(item_sel_opcao)
                            item = matches_para_usar.iloc[idx_match]
                        else: item = matches_para_usar.iloc[0]

                        lote_auto = str(item.get('lote', '')).strip()
                        ativo_auto = str(item.get('ativo', '')).strip()
                        id_est_limpo = str(item.get('id_estoque_fisico', '')).strip() or extrair_id_estoque_do_nome(inventario_selected_obj['nome'])
                        desc_est_limpo = str(item.get('desc_estoque_fisico', '')).strip() or MAPA_ESTOQUES_DESC.get(id_est_limpo, 'ESTOQUE FÍSICO')

                        c1, c2, c3 = st.columns(3)
                        c1.metric("Cód. Produto", str(item['cod_produto']))
                        c2.metric("Lote", lote_auto or "—")
                        c3.metric("Ativo", ativo_auto or "—")
                        
                        st.markdown(f'<div class="caixa-descricao-produto">📦 Descrição: {item["desc_produto"]}</div>', unsafe_allow_html=True)
                        st.info(f"📍 **Estoque Físico:** {desc_est_limpo} (ID: {id_est_limpo}) | **Unidade:** {item.get('unid_medida', 'UN')} | **Status:** {status_pasta_atual}")
                        
                        with st.form("form_lancar_qtd", clear_on_submit=True):
                            qtd_fisica = st.number_input("📦 Quantidade Contada Fisicamente:", min_value=0, step=1, value=0)
                            confirma_zero = st.checkbox("⚠️ Marque se este item REALMENTE NÃO EXISTE no estoque (Saldo Zero)")
                            obs = st.text_input("Observação (opcional)")
                            
                            if st.form_submit_button("⚡ Salvar na RAM (Instantâneo)", type="primary", use_container_width=True):
                                if qtd_fisica == 0 and not confirma_zero: st.error("⚠️ Para salvar quantidade 0, marque a confirmação amarela!")
                                else:
                                    qtd_sys = int(pd.to_numeric(item.get('qtd_estoque', 0), errors='coerce') or 0)
                                    dif = qtd_fisica - qtd_sys
                                    fase_gravar = "2a Contagem Concluida" if status_pasta_atual == "2a Contagem" else status_pasta_atual
                                    data_hora_agora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                                    st.session_state.buffer_ram_contagens.append({
                                        'inventario_id': id_pasta_limpo_base,
                                        'id_estoque': id_est_limpo,
                                        'desc_estoque': desc_est_limpo,
                                        'cod_produto': str(item['cod_produto']),
                                        'desc_produto': str(item['desc_produto']),
                                        'unid_medida': str(item.get('unid_medida', 'UN')),
                                        'qtd_sistema': qtd_sys,
                                        'qtd_contada': qtd_fisica,
                                        'diferenca': dif,
                                        'ativo': ativo_auto,
                                        'observacao': obs,
                                        'operador': st.session_state.operador,
                                        'data_hora': data_hora_agora,
                                        'lote': lote_auto,
                                        'fase_contagem': fase_gravar
                                    })

                                    st.session_state.contador_reset += 1
                                    st.toast("⚡ Salvo na RAM instantaneamente!", icon="✅")
                                    st.rerun()

    # --- ABA 2: LANÇAMENTOS E ESPELHO BASE ---
    with aba_lancamentos:
        sub_aba1, sub_aba2 = st.tabs(["📋 Meus Lançamentos Nesta Pasta", "📄 Espelho Base do Saldo (Status Visual)"])
        with sub_aba1:
            df_cnts_todas = ler_aba("contagens")
            df_minhas = df_cnts_todas[df_cnts_todas['inventario_id'].astype(str) == id_pasta_limpo_base] if not df_cnts_todas.empty else pd.DataFrame()
            
            if st.session_state.buffer_ram_contagens:
                df_ram = pd.DataFrame(st.session_state.buffer_ram_contagens)
                df_ram_pasta = df_ram[df_ram['inventario_id'].astype(str) == id_pasta_limpo_base] if not df_ram.empty else pd.DataFrame()
                if not df_ram_pasta.empty:
                    df_minhas = pd.concat([df_ram_pasta, df_minhas], ignore_index=True)

            if not df_minhas.empty:
                m1, m2, m3 = st.columns(3)
                m1.metric("Lançamentos Totais", len(df_minhas))
                m2.metric("Com Divergência", len(df_minhas[pd.to_numeric(df_minhas['diferenca'], errors='coerce') != 0]))
                m3.metric("Total de Peças Contadas", int(pd.to_numeric(df_minhas['qtd_contada'], errors='coerce').sum()))
                st.download_button("📥 Exportar Lançamentos Filtrados para Excel", converter_para_excel(df_minhas), file_name=f"contagem_pasta_{id_pasta_limpo_base}.xlsx")
                st.dataframe(df_minhas[['cod_produto', 'desc_produto', 'desc_estoque', 'qtd_sistema', 'qtd_contada', 'diferenca', 'ativo', 'lote', 'observacao', 'operador', 'data_hora', 'fase_contagem']], use_container_width=True, hide_index=True)
            else: st.info("Nenhum lançamento registrado nesta pasta.")
            
        with sub_aba2:
            if not base_sistema_atual.empty:
                df_cnts_todas = ler_aba("contagens")
                df_cnts_p = df_cnts_todas[df_cnts_todas['inventario_id'].astype(str) == id_pasta_limpo_base] if not df_cnts_todas.empty else pd.DataFrame()
                
                mapa_contados = {}
                if not df_cnts_p.empty:
                    for _, r in df_cnts_p.iterrows():
                        key = f"{str(r.get('cod_produto','')).upper().strip()}_{str(r.get('lote','')).upper().strip()}_{str(r.get('ativo','')).upper().strip()}"
                        mapa_contados[key] = str(r.get('operador', 'Operador'))
                
                for r_ram in st.session_state.buffer_ram_contagens:
                    if str(r_ram['inventario_id']) == id_pasta_limpo_base:
                        key = f"{str(r_ram['cod_produto']).upper().strip()}_{str(r_ram['lote']).upper().strip()}_{str(r_ram['ativo']).upper().strip()}"
                        mapa_contados[key] = r_ram['operador']

                def obter_status(row):
                    c = str(row.get('cod_produto','')).upper().strip()
                    l = str(row.get('lote','')).upper().strip() if pd.notna(row.get('lote')) and str(row.get('lote')).lower()!='nan' else ""
                    a = str(row.get('ativo','')).upper().strip() if pd.notna(row.get('ativo')) and str(row.get('ativo')).lower()!='nan' else ""
                    key = f"{c}_{l}_{a}"
                    if key in mapa_contados: return f"🟩 Contabilizado por ({mapa_contados[key]})"
                    return "🟥 Não Contado"
                
                df_espelho = base_sistema_atual.copy()
                df_espelho['Status de Contagem'] = df_espelho.apply(obter_status, axis=1)
                st.dataframe(df_espelho[['Status de Contagem', 'cod_produto', 'desc_produto', 'desc_estoque_fisico', 'unid_medida', 'qtd_estoque', 'id_estoque_fisico', 'lote', 'ativo']], use_container_width=True, hide_index=True)
            else: st.info("Nenhuma base carregada.")

    # --- ABA 3: DESEMPENHO E ACURACIDADE ---
    with aba_desempenho:
        sub_d1, sub_d2 = st.tabs(["🔴 Desempenho & Prazos por Estoque", "📈 Acuracidade Auditada pelo Supervisor"])
        with sub_d1:
            st.subheader("🏆 Status de Atualização dos Estoques Físicos")
            mapa_datas = {}
            df_ult_historico = ler_aba("ultima_contagem_estoques")
            df_cnts_todas = ler_aba("contagens")
            
            if not df_ult_historico.empty:
                for _, r_h in df_ult_historico.iterrows(): mapa_datas[str(r_h.get('id_estoque','')).strip()] = str(r_h.get('ultima_data',''))
            if not df_cnts_todas.empty:
                for _, r_u in df_cnts_todas.iterrows():
                    id_e = str(r_u.get('id_estoque','')).strip()
                    if id_e and id_e not in mapa_datas: mapa_datas[id_e] = str(r_u.get('data_hora',''))

            linhas_desempenho, hoje = [], datetime.datetime.now()
            b_count, a_count, c_count = 0, 0, 0
            
            for est in LISTA_ESTOQUES_FIXA:
                est_id = str(est["id"]).strip()
                u_data = mapa_datas.get(est_id, None)
                if u_data:
                    try:
                        str_dt = str(u_data).split(".")[0]
                        if len(str_dt) == 10: str_dt += " 12:00:00"
                        dt = datetime.datetime.strptime(str_dt, "%Y-%m-%d %H:%M:%S")
                        dias, dt_fmt = (hoje - dt).days, dt.strftime("%d/%m/%Y %H:%M")
                    except Exception: dias, dt_fmt = 999, "Sem histórico"
                else: dias, dt_fmt = 999, "Nunca Contado"
                
                if dias <= 7: status, b_count = "🟢 Em Dia", b_count + 1
                elif dias <= 14: status, a_count = "🟡 Necessário Auditar", a_count + 1
                else: status, c_count = "🔴 Crítico (+2 semanas)", c_count + 1
                    
                linhas_desempenho.append({"Id. Estoque": est_id, "Descrição do Estoque Físico": est["desc"], "Última Contagem": dt_fmt, "Dias Sem Contar": dias if dias != 999 else 999, "Criticidade": status})

            df_desempenho_full = pd.DataFrame(linhas_desempenho)
            k1, k2, k3 = st.columns(3)
            k1.metric("🟢 Em Dia (até 7 dias)", b_count)
            k2.metric("🟡 Necessário Auditar (8 a 14 dias)", a_count)
            k3.metric("🔴 Crítico (+2 semanas)", c_count)
            
            col_f_stat, col_f_busca = st.columns([1, 1])
            filtro_criticidade = col_f_stat.selectbox("🎯 Filtrar por Criticidade:", ["Todos os Estoques", "🟢 Em Dia", "🟡 Necessário Auditar", "🔴 Crítico (+2 semanas)"])
            busca_estoque = col_f_busca.text_input("🔍 Pesquisar Estoque por Código ou Nome:")

            df_exibir = df_desempenho_full.copy()
            if filtro_criticidade != "Todos os Estoques": df_exibir = df_exibir[df_exibir['Criticidade'] == filtro_criticidade]
            if busca_estoque.strip():
                termo = busca_estoque.strip().lower()
                df_exibir = df_exibir[df_exibir['Id. Estoque'].str.lower().str.contains(termo) | df_exibir['Descrição do Estoque Físico'].str.lower().str.contains(termo)]

            df_exibir['Dias Sem Contar'] = df_exibir['Dias Sem Contar'].apply(lambda x: "—" if x == 999 else x)
            st.dataframe(df_exibir.sort_values(by=['Criticidade', 'Id. Estoque'], ascending=[False, True]), use_container_width=True, hide_index=True)

        with sub_d2:
            st.subheader("📈 Tabela de Acuracidade Geral por Depósito")
            df_auds = ler_aba("auditorias_supervisor")
            if df_auds.empty: st.info("💡 Nenhuma amostragem coletada pelo supervisor.")
            else:
                linhas_acu = []
                for dep_id, grp in df_auds.groupby('id_estoque'):
                    tot = len(grp)
                    difs_num = pd.to_numeric(grp['diferenca'], errors='coerce')
                    p_s = (len(grp[difs_num == 0])/tot)*100
                    p_e = (len(grp[grp['etiqueta_correta'] == "Sim"])/tot)*100
                    p_l = (len(grp[grp['localizacao_correta'] == "Sim"])/tot)*100
                    linhas_acu.append({"CÓDIGO ESTOQUE": dep_id, "DESCRIÇÃO DO ESTOQUE": grp.iloc[0].get('desc_estoque', 'Não Informado'), "ACURACIDADE SALDO": f"{'🟢' if p_s==100 else '🔴'} {p_s:.1f}%", "ACURACIDADE ETIQUETAS": f"{'🟢' if p_e==100 else '🔴'} {p_e:.1f}%", "ACURACIDADE LOCALIZAÇÃO": f"{'🟢' if p_l==100 else '🔴'} {p_l:.1f}%", "ITENS AUDITADOS": tot, "ÚLTIMA AUDITORIA": str(grp.iloc[0].get('data_hora', '')).split(" ")[0]})
                st.dataframe(pd.DataFrame(linhas_acu), use_container_width=True, hide_index=True)

            st.markdown("---")
            if not df_inventarios_sup.empty:
                for _, inv_s in df_inventarios_sup.iterrows():
                    df_hist_sup = df_auds[df_auds['inventario_id'].astype(str) == str(inv_s['id'])] if not df_auds.empty else pd.DataFrame()
                    c_exp, c_del = st.columns([8, 2])
                    with c_exp:
                        with st.expander(f"📁 Pasta {inv_s['id']} – {inv_s['nome']} | Data: {inv_s['data']} | Status: {inv_s['status']} ({len(df_hist_sup)} itens auditados)"):
                            if not df_hist_sup.empty:
                                st.download_button("📥 Exportar Esta Auditoria para Excel", converter_para_excel(df_hist_sup), file_name=f"auditoria_{inv_s['id']}.xlsx", key=f"dl_sup_acu_{inv_s['id']}")
                                st.dataframe(df_hist_sup, use_container_width=True, hide_index=True)
                            else: st.info("Nenhum item auditado nesta pasta.")
                    with c_del:
                        if eh_supervisor and st.button("🗑️ Excluir Pasta", key=f"del_sup_f_{inv_s['id']}", use_container_width=True):
                            df_inv_sup_up = df_inventarios_sup[df_inventarios_sup['id'].astype(str) != str(inv_s['id'])]
                            atualizar_aba_completa("inventarios_supervisor", df_inv_sup_up)
                            limpar_cache_aplicacao()
                            st.rerun()

    # --- ABA 4: HISTÓRICO GERAL ---
    with aba_historico:
        st.title("📁 Arquivo Geral de Movimentações")
        if df_inventarios.empty or 'id' not in df_inventarios.columns: 
            st.info("Nenhum inventário registrado.")
        else:
            df_inv_ordenados = df_inventarios.copy()
            df_inv_ordenados['id_num'] = pd.to_numeric(df_inv_ordenados['id'].astype(str).str.replace('#', '', regex=False), errors='coerce').fillna(0)
            df_inv_ordenados = df_inv_ordenados.sort_values(by=['id_num'], ascending=False)

            itens_por_pagina = 10
            total_itens = len(df_inv_ordenados)
            total_paginas = max(1, (total_itens + itens_por_pagina - 1) // itens_por_pagina)

            if st.session_state.pagina_historico > total_paginas:
                st.session_state.pagina_historico = total_paginas

            inicio_idx = (st.session_state.pagina_historico - 1) * itens_por_pagina
            fatia_pastas = df_inv_ordenados.iloc[inicio_idx:inicio_idx + itens_por_pagina]

            for idx, inv in fatia_pastas.iterrows():
                id_proc = str(inv['id']).replace('#','').strip()
                df_cnts_todas = ler_aba("contagens")
                
                df_h_drive = df_cnts_todas[df_cnts_todas['inventario_id'].astype(str) == id_proc] if not df_cnts_todas.empty else pd.DataFrame()
                df_h_ram = pd.DataFrame([c for c in st.session_state.buffer_ram_contagens if str(c['inventario_id']) == id_proc])
                
                df_h = pd.concat([df_h_ram, df_h_drive], ignore_index=True)
                
                tot_reg = len(df_h)
                acu_reg = inv.get('acuracidade_final', '—')

                c_exp, c_del = st.columns([8, 2])
                with c_exp:
                    with st.expander(f"📁 Pasta {inv['id']} - {inv['nome']} | Data: {inv['data']} | Status: {inv['status']} | Acuracidade: {acu_reg} ({tot_reg} itens contados)"):
                        if not df_h.empty:
                            st.download_button("📥 Baixar Planilha (.xlsx)", converter_para_excel(df_h), file_name=f"inventario_{id_proc}.xlsx", key=f"dl_hist_{id_proc}")
                            st.dataframe(df_h, use_container_width=True, hide_index=True)
                        else: st.info("Nenhum lançamento registrado nesta pasta.")
                with c_del:
                    if eh_supervisor and st.button("🗑️ Excluir Pasta", key=f"del_hist_inv_{inv['id']}", use_container_width=True):
                        st.session_state.buffer_ram_inventarios = [i for i in st.session_state.buffer_ram_inventarios if str(i['id']) != str(inv['id'])]
                        if not df_inv_drive.empty:
                            df_inv_drive_att = df_inv_drive[df_inv_drive['id'].astype(str) != str(inv['id'])]
                            atualizar_aba_completa("inventarios", df_inv_drive_att)
                        limpar_cache_aplicacao()
                        st.rerun()

            st.markdown("---")
            c_pag1, c_pag2, c_pag3 = st.columns([2, 3, 2])
            with c_pag1:
                if st.button("◀ Anterior", disabled=(st.session_state.pagina_historico <= 1), use_container_width=True):
                    st.session_state.pagina_historico -= 1
                    st.rerun()
            with c_pag2:
                st.markdown(f"<h5 style='text-align: center;'>Página {st.session_state.pagina_historico} de {total_paginas} ({total_itens} pastas)</h5>", unsafe_allow_html=True)
            with c_pag3:
                if st.button("Próxima ▶", disabled=(st.session_state.pagina_historico >= total_paginas), use_container_width=True):
                    st.session_state.pagina_historico += 1
                    st.rerun()

    # --- ABA 5: GESTÃO ADM ---
    if eh_supervisor and aba_adm is not None:
        with aba_adm:
            st.title("⚙️ Módulo de Gestão do Administrador (Google Drive)")
            opcao_adm = st.selectbox("Escolha o Módulo de Ação:", ["🚨 Liberar / Encerrar Divergências", "🔬 Auditoria Amostral (Supervisor)", "📊 Relatório Consolidado (Excel Gerencial)", "👥 Gestão de Usuários & Senhas"])
            st.markdown("---")

            if opcao_adm == "🚨 Liberar / Encerrar Divergências":
                st.subheader("Tratamento de Erros de Contagem da Equipe")
                df_cnts = ler_aba("contagens")
                if not df_cnts.empty:
                    df_cnts['diferenca_num'] = pd.to_numeric(df_cnts['diferenca'], errors='coerce').fillna(0)
                    df_divs = df_cnts[(df_cnts['diferenca_num'] != 0) & (~df_cnts['fase_contagem'].astype(str).isin(['2a Contagem', 'Encerrado com Divergencia']))]
                    
                    if df_divs.empty:
                        st.success("🎉 Nenhuma divergência pendente no momento!")
                    else:
                        opcoes_div = [f"Item ID #{idx} - {r['cod_produto']} - {r['desc_produto']} (Dif: {r['diferenca']})" for idx, r in df_divs.iterrows()]
                        sel_div = st.selectbox("Selecione o Item com Divergência:", opcoes_div)
                        idx_sel = int(sel_div.split("Item ID #")[1].split(" - ")[0])
                        row_target = df_divs.loc[idx_sel]
                        
                        st.info(f"📊 Sistema: {row_target['qtd_sistema']} | Contado: {row_target['qtd_contada']} | Dif: {row_target['diferenca']}")
                        justificativa_adm = st.text_input("Justificativa:", value="Divergente")
                        
                        col_act1, col_act2 = st.columns(2)
                        with col_act1:
                            if st.button("🚨 Abrir 2ª Contagem", type="primary", use_container_width=True):
                                df_cnts.loc[idx_sel, 'fase_contagem'] = '2a Contagem'
                                df_cnts.loc[idx_sel, 'observacao'] = f"ADM ({st.session_state.operador}): [LIBERADO 2ª CONTAGEM] - {justificativa_adm.strip()}"
                                atualizar_aba_completa("contagens", df_cnts.drop(columns=['diferenca_num']))
                                
                                id_pasta_target = str(row_target['inventario_id']).replace('#', '').strip()
                                df_inv_all = ler_aba("inventarios")
                                if not df_inv_all.empty:
                                    mascara_pasta = df_inv_all['id'].astype(str).str.replace('#', '').str.strip() == id_pasta_target
                                    df_inv_all.loc[mascara_pasta, 'status'] = '2a Contagem'
                                    sincronizar_upsert_inventarios(df_inv_all)

                                for inv_r in st.session_state.buffer_ram_inventarios:
                                    if str(inv_r['id']).replace('#', '').strip() == id_pasta_target:
                                        inv_r['status'] = '2a Contagem'

                                st.success("✅ Item liberado e Pasta Reaberta para 2ª Contagem!")
                                st.rerun()

                        with col_act2:
                            if st.button("🔒 Finalizar e Manter Divergência", use_container_width=True):
                                df_cnts.loc[idx_sel, 'fase_contagem'] = 'Encerrado com Divergencia'
                                df_cnts.loc[idx_sel, 'observacao'] = f"ADM ({st.session_state.operador}): [ENCERRADO COM DIVERGÊNCIA] - {justificativa_adm.strip()}"
                                atualizar_aba_completa("contagens", df_cnts.drop(columns=['diferenca_num']))
                                st.success("✅ Divergência encerrada!")
                                st.rerun()
                else:
                    st.info("Nenhuma contagem no sistema.")

            elif opcao_adm == "🔬 Auditoria Amostral (Supervisor)":
                st.subheader("Módulo de Auditoria Amostral Própria")
                if df_inventarios_sup.empty: id_sup_act, inv_sup_obj = None, None
                else:
                    sel_s = st.selectbox("Selecione a Pasta de Auditoria Ativa:", [f"{r['id']} – {r['nome']} ({r['status']})" for _, r in df_inventarios_sup.iterrows()], key="sb_sup_active")
                    id_sup_act = sel_s.split(" – ")[0]
                    inv_sup_obj = df_inventarios_sup[df_inventarios_sup['id'].astype(str) == id_sup_act].iloc[0]

                if inv_sup_obj is not None and str(inv_sup_obj['status']) == "Aberto":
                    if st.button("🔒 Fechar Esta Pasta de Auditoria", type="primary"):
                        df_inventarios_sup.loc[df_inventarios_sup['id'].astype(str) == str(id_sup_act), 'status'] = 'Fechado'
                        atualizar_aba_completa("inventarios_supervisor", df_inventarios_sup)
                        limpar_cache_aplicacao()
                        st.rerun()

                with st.expander("➕ Nova Pasta de Auditoria do Supervisor"):
                    with st.form("form_sup_new"):
                        nom_s = st.text_input("Nome da Pasta Amostral")
                        if st.form_submit_button("Criar Pasta Supervisor", type="primary") and nom_s:
                            m_id_s = len(df_inventarios_sup) + 1
                            df_novo_sup = pd.DataFrame([{
                                "id": f"SUP-#{m_id_s}",
                                "nome": nom_s.strip(),
                                "data": datetime.date.today().strftime("%Y-%m-%d"),
                                "status": "Aberto"
                            }])
                            salvar_lote_aba("inventarios_supervisor", df_novo_sup)
                            limpar_cache_aplicacao()
                            st.rerun()

                arq_sup = st.file_uploader("Suba a planilha Excel de amostras (.xlsx)", type=["xlsx"], key="up_excel_sup")
                if arq_sup is not None and id_sup_act:
                    try:
                        st.session_state.bases_supervisor_por_inv[id_sup_act] = pd.read_excel(arq_sup)
                        st.success("✅ Planilha anexa com sucesso!")
                    except Exception as e: st.error(f"Erro: {e}")

                base_sup_curr = st.session_state.bases_supervisor_por_inv.get(id_sup_act, None)

                if id_sup_act and base_sup_curr is not None and str(inv_sup_obj['status']) == "Aberto":
                    cols_sup = list(base_sup_curr.columns)
                    def mapear_col_s(opcoes, idx_padrao):
                        for op in opcoes:
                            for c in cols_sup:
                                if op.lower().replace(" ", "").replace(".", "") in str(c).lower().replace(" ", "").replace(".", ""): return c
                        return cols_sup[idx_padrao] if idx_padrao < len(cols_sup) else cols_sup[0]

                    col_cod_s, col_desc_s = mapear_col_s(['códproduto', 'codproduto', 'codigo'], 0), mapear_col_s(['descproduto', 'descricao'], 1)
                    col_local_s, col_qtd_s, col_id_est_s = mapear_col_s(['descestoquefisico', 'localizacao'], 2), mapear_col_s(['qtdestoque', 'quantidade'], -1), mapear_col_s(['idestoquefísico', 'idestoque'], 0)

                    df_auds_todas = ler_aba("auditorias_supervisor")
                    df_ja_auditados = df_auds_todas[df_auds_todas['inventario_id'].astype(str) == str(id_sup_act)] if not df_auds_todas.empty else pd.DataFrame()
                    cods_ja_auditados = set(df_ja_auditados['cod_produto'].astype(str).str.upper().str.strip().tolist()) if not df_ja_auditados.empty else set()

                    tot_amostra = len(base_sup_curr)
                    st.progress(min(1.0, len(cods_ja_auditados) / tot_amostra) if tot_amostra > 0 else 0.0)
                    st.dataframe(base_sup_curr, use_container_width=True, hide_index=True)

                    base_pendente = base_sup_curr[~base_sup_curr[col_cod_s].astype(str).str.upper().str.strip().isin(cods_ja_auditados)]

                    col_bip, col_select = st.columns([1, 1])
                    bip_sup = col_bip.text_input("Bipe o código com leitor (opcional):", key=f"bip_sup_{st.session_state.contador_reset_sup}")
                    item_combo_sup = col_select.selectbox("Ou selecione o material da lista:", [f"{r[col_cod_s]} - {r[col_desc_s]}" for _, r in base_pendente.iterrows()] if not base_pendente.empty else [], key="combo_sup_amostra") if not base_pendente.empty else None

                    cod_sup_clean = str(bip_sup).upper().strip().split(" - ")[-1].strip() if bip_sup else (str(item_combo_sup).split(" - ")[0].strip() if item_combo_sup else "")

                    if cod_sup_clean:
                        match_s = base_sup_curr[base_sup_curr[col_cod_s].astype(str).str.upper().str.strip() == cod_sup_clean]
                        if not match_s.empty:
                            row_s = match_s.iloc[0]
                            st.info(f"📦 **Auditando Item:** {cod_sup_clean} - {row_s[col_desc_s]}")
                            with st.form("form_auditar_item_completo", clear_on_submit=True):
                                c_f1, c_f2, c_f3 = st.columns(3)
                                q_aud = c_f1.number_input("Quantidade Real Encontrada Fisicamente:", min_value=0, step=1, value=0)
                                e_ok = c_f2.selectbox("A Etiqueta Física está Correta?", ["Sim", "Não"])
                                l_ok = c_f3.selectbox("O Endereçamento/Localização está Correto?", ["Sim", "Não"])
                                at_sup = st.text_input("Número do Ativo (Opcional)")
                                if st.form_submit_button("💾 Salvar Auditoria do Item", type="primary", use_container_width=True):
                                    qtd_sys = int(pd.to_numeric(row_s[col_qtd_s], errors='coerce') or 0)
                                    df_nova_aud = pd.DataFrame([{
                                        "inventario_id": id_sup_act,
                                        "id_estoque": str(row_s[col_id_est_s]).strip() if col_id_est_s in base_sup_curr.columns else "",
                                        "desc_estoque": str(row_s[col_local_s]) if col_local_s in base_sup_curr.columns else "Não Informado",
                                        "cod_produto": cod_sup_clean,
                                        "desc_produto": str(row_s[col_desc_s]),
                                        "qtd_sistema": qtd_sys,
                                        "qtd_auditada": q_aud,
                                        "diferenca": q_aud - qtd_sys,
                                        "etiqueta_correta": e_ok,
                                        "localizacao_correta": l_ok,
                                        "supervisor": st.session_state.operador,
                                        "data_hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                        "recontagem_3": "Não",
                                        "ativo": at_sup.strip().upper()
                                    }])
                                    salvar_lote_aba("auditorias_supervisor", df_nova_aud)
                                    limpar_cache_aplicacao()
                                    st.success("✅ Auditoria registrada!")
                                    st.session_state.contador_reset_sup += 1
                                    st.rerun()

            elif opcao_adm == "📊 Relatório Consolidado (Excel Gerencial)":
                st.subheader("Gerar Planilha Gerencial por Período (Múltiplas Abas)")
                
                hoje = datetime.date.today()
                quinze_dias_atras = hoje - datetime.timedelta(days=15)
                
                col_d1, col_d2 = st.columns(2)
                with col_d1:
                    data_inicio = st.date_input("📅 Data Inicial:", value=quinze_dias_atras, format="DD/MM/YYYY")
                with col_d2:
                    data_fim = st.date_input("📅 Data Final:", value=hoje, format="DD/MM/YYYY")
                
                str_ini = data_inicio.strftime("%Y-%m-%d")
                str_fim = data_fim.strftime("%Y-%m-%d")
                
                df_pastas_todas = ler_aba("inventarios")
                if not df_pastas_todas.empty:
                    df_pastas_todas['data_dt'] = pd.to_datetime(df_pastas_todas['data'], errors='coerce')
                    df_pastas_periodo = df_pastas_todas[(df_pastas_todas['data_dt'] >= pd.to_datetime(str_ini)) & (df_pastas_todas['data_dt'] <= pd.to_datetime(str_fim))]
                else: df_pastas_periodo = pd.DataFrame()
                
                if df_pastas_periodo.empty:
                    st.info(f"ℹ️ Nenhum inventário registrado entre **{data_inicio.strftime('%d/%m/%Y')}** e **{data_fim.strftime('%d/%m/%Y')}**.")
                else:
                    st.success(f"📌 Foram encontrados **{len(df_pastas_periodo)}** inventários no período de {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}.")
                    
                    ids_pastas = [str(r['id']).replace('#', '').strip() for _, r in df_pastas_periodo.iterrows()]
                    
                    if st.button("🚀 Gerar Excel Consolidado do Período", type="primary", use_container_width=True):
                        df_c_todas = ler_aba("contagens")
                        if not df_c_todas.empty:
                            df_c_fil = df_c_todas[df_c_todas['inventario_id'].astype(str).str.replace('#', '').isin(ids_pastas)]
                        else: df_c_fil = pd.DataFrame()
                        
                        if df_c_fil.empty:
                            st.warning("⚠️ Os inventários do período foram criados, mas ainda não possuem nenhum lançamento de contagem registrado.")
                        else:
                            bytes_ex = gerar_relatorio_consolidado_excel(df_c_fil, LISTA_ESTOQUES_FIXA)
                            st.download_button(
                                label=f"📥 Clique para Baixar o Relatório ({data_inicio.strftime('%d-%m-%Y')} a {data_fim.strftime('%d-%m-%Y')}).xlsx", 
                                data=bytes_ex, 
                                file_name=f"Relatorio_Gerencial_{str_ini}_a_{str_fim}.xlsx", 
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                                use_container_width=True
                            )

            elif opcao_adm == "👥 Gestão de Usuários & Senhas":
                st.subheader("Gerenciamento de Colaboradores e Perfis")
                df_usrs = ler_aba("usuarios")
                if df_usrs.empty: st.info("Nenhum usuário cadastrado.")
                else:
                    st.dataframe(df_usrs[['nome', 'cpf', 'email', 'perfil']], use_container_width=True, hide_index=True)
                    c_u1, c_u2 = st.columns(2)
                    with c_u1:
                        u_sel = st.selectbox("Escolha o Colaborador:", [f"{idx} - {r['nome']}" for idx, r in df_usrs.iterrows()])
                        idx_usr = int(u_sel.split(" - ")[0])
                        n_senha, n_perfil = st.text_input("Nova Senha", type="password"), st.selectbox("Nível de Acesso:", ["Almoxarife", "Administrador"])
                        if st.button("🔄 Atualizar Dados do Usuário", type="primary", use_container_width=True):
                            if n_senha.strip(): df_usrs.loc[idx_usr, 'senha'] = n_senha.strip()
                            df_usrs.loc[idx_usr, 'perfil'] = n_perfil
                            atualizar_aba_completa("usuarios", df_usrs)
                            st.success("✅ Atualizado!")
                            st.rerun()
                    with c_u2:
                        u_del = st.selectbox("Remover Colaborador:", [f"{idx} - {r['nome']}" for idx, r in df_usrs.iterrows()], key="sb_del")
                        idx_del = int(u_del.split(" - ")[0])
                        if st.button("❌ Confirmar Exclusão", type="primary", use_container_width=True):
                            df_usrs = df_usrs.drop(index=idx_del)
                            atualizar_aba_completa("usuarios", df_usrs)
                            st.success("✅ Usuário removido!")
                            st.rerun()
